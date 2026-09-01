# -*- coding: utf-8 -*-
"""Comprueba que el .xlsx que genera el sitio se pueda abrir de verdad.

    python codigo/verificar_xlsx.py

El escritor de xlsx de tpl_tail.js arma el archivo a mano (un xlsx es un ZIP
con unos XML dentro) sin ninguna libreria externa. Eso puede quedar
sutilmente mal -- un CRC-32 equivocado, un offset corrido en el directorio
central, un caracter de control que invalida el XML -- de formas que no se
ven hasta que alguien intenta abrir el archivo en Excel y le dice que esta
danado.

Esta prueba cierra ese hueco de punta a punta:

  1. Corre codigo/verificar_xlsx.mjs (Node), que extrae las funciones REALES
     del HTML ya generado -- no una copia -- y exporta una muestra de UDS.
  2. Abre el resultado con openpyxl (el mismo lector que usa el pipeline para
     leer la hoja fuente) y compara celda por celda contra lo que se esperaba.

Requiere Node.js. Sale con codigo 1 si algo falla.
"""
import json
import os
import subprocess
import sys
import tempfile
import zipfile

import openpyxl

D = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(D)
SITIO = os.path.join(RAIZ, "DONDE LO UBICO - Cercania UDS Antioquia.html")

if not os.path.exists(SITIO):
    raise SystemExit("Falta el sitio generado. Corre primero: python codigo/generar_sitio.py")

tmp = tempfile.mkdtemp(prefix="verif_xlsx_")
destino = os.path.join(tmp, "prueba.xlsx")

try:
    r = subprocess.run(
        ["node", os.path.join(D, "verificar_xlsx.mjs"), destino],
        capture_output=True, text=True,
    )
except FileNotFoundError:
    raise SystemExit(
        "No se encontro Node.js, necesario para ejecutar el escritor de xlsx\n"
        "tal como corre en el navegador. Instalalo, o salta esta prueba (el\n"
        "resto de verificaciones estan en codigo/verificar.py)."
    )
if r.returncode != 0:
    print(r.stdout)
    print(r.stderr, file=sys.stderr)
    raise SystemExit("Fallo al generar el .xlsx de prueba.")
print(r.stdout.strip())

fallos = []


def check(cond, nombre, detalle=""):
    if cond:
        print("  OK   %s" % nombre)
    else:
        print("  FALLA %s -- %s" % (nombre, detalle))
        fallos.append(nombre)


print("\n=== El archivo es un ZIP integro ===")
check(zipfile.is_zipfile(destino), "es un ZIP valido")
z = zipfile.ZipFile(destino)
check(z.testzip() is None, "todos los CRC-32 cuadran", "entrada corrupta: %s" % z.testzip())
partes = set(z.namelist())
esperadas = {
    "[Content_Types].xml", "_rels/.rels", "xl/workbook.xml",
    "xl/_rels/workbook.xml.rels", "xl/worksheets/sheet1.xml",
}
check(esperadas <= partes, "estan las 5 partes que exige el formato",
      "faltan: %s" % (esperadas - partes))

print("\n=== Un lector de Excel real lo abre ===")
wb = openpyxl.load_workbook(destino)
check(wb.sheetnames == ["CUENTAME"], "la hoja se llama CUENTAME",
      "se encontro %s" % wb.sheetnames)
ws = wb["CUENTAME"]

with open(destino + ".esperado.json", encoding="utf-8") as f:
    esperado = [["" if v is None else v for v in fila] for fila in json.load(f)]
leido = [["" if c is None else c for c in fila] for fila in ws.iter_rows(values_only=True)]

check(len(leido) == len(esperado), "mismo numero de filas",
      "esperadas %d, leidas %d" % (len(esperado), len(leido)))
check(ws.max_column == len(esperado[0]), "mismo numero de columnas",
      "esperadas %d, leidas %d" % (len(esperado[0]), ws.max_column))

print("\n=== El contenido coincide celda por celda ===")
distintas = []
for i, (a, b) in enumerate(zip(esperado, leido)):
    for j, (x, y) in enumerate(zip(a, b)):
        if str(x) != str(y):
            distintas.append((i + 1, j + 1, x, y))
check(not distintas, "ninguna celda difiere",
      "%d distintas, ej %s" % (len(distintas), distintas[:3]))
print("       (%d celdas comparadas)" % sum(len(f) for f in esperado))

z.close()
wb.close()
for f in (destino, destino + ".esperado.json"):
    try:
        os.remove(f)
    except OSError:
        pass
try:
    os.rmdir(tmp)
except OSError:
    pass

print("\n" + "=" * 60)
if fallos:
    raise SystemExit("RESULTADO: %d PRUEBA(S) FALLARON -> %s" % (len(fallos), fallos))
print("RESULTADO: el .xlsx que genera el sitio se abre y su contenido es exacto")
