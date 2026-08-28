# -*- coding: utf-8 -*-
"""
Extrae de "BD ANALISIS COBERTURA v21.xlsm" (hoja CUENTAME) todas las UDS
ACTIVAS del departamento de Antioquia, las cruza con los poligonos reales
de "antioquia_con_comunas v5.geojson" y arma el dataset que consume
"DONDE LO UBICO - Cercania UDS Antioquia.html" (const D = __DATOS__;).

Por que se parsean las coordenadas de CUENTAME en vez de usar la hoja
"COORDENADAS UDS": esa hoja solo cruza por Codigo UDS contra CUENTAME en
~75% de los casos (3230/4329) porque tiene mas filas (contratos viejos)
que no corresponden 1 a 1. Las columnas "Latitud UDS"/"Longitud UDS" de
CUENTAME en cambio traen grados/minutos/segundos como texto (ej.
6.11'13.9" N) para el 100% de las UDS activas (4328/4328) -- parsearlas
directamente da cobertura completa sin depender del cruce.
"""
import datetime
import json
import os
import re
import unicodedata

import openpyxl

D = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(D)
FUENTE = os.path.join(RAIZ, "fuente", "BD ANALISIS COBERTURA v21.xlsm")
FUENTE_GEO = os.path.join(RAIZ, "fuente", "antioquia_con_comunas v5.geojson")
FUENTE_CASCOS = os.path.join(RAIZ, "fuente", "COORDENADAS MUNICIPIOS.xlsx")
OUT = os.path.join(RAIZ, "recursos", "datos_cercania.json")


def norm(s):
    """Normaliza un nombre de municipio para cruzar geojson <-> Excel: sin
    tildes/enye, solo A-Z0-9, mayusculas -- ninguna de las dos fuentes usa
    la misma convencion de acentos/espacios."""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Z0-9]", "", s.upper())


# 2 municipios donde el nombre oficial completo del geojson no coincide con
# el nombre corto que trae "Municipio UDS" en el Excel -- unicos 2 casos
# tras normalizar tildes/espacios (verificado 2026-08-27: los demas casos
# que parecian no cruzar, como EL PEÑOL/EL RETIRO/SAN VICENTE o el
# "MPIO_NOMBR": "rionegro" mal cargado en el feature de Zaragoza, se
# resuelven solos usando la propiedad "name" del geojson en vez de
# "MPIO_NOMBR", que es la que trae el nombre corto/correcto).
ALIAS_MUNICIPIO = {
    norm("SAN ANDRES DE CUERQUIA"): norm("SAN ANDRES"),
    norm("SAN PEDRO DE LOS MILAGROS"): norm("SAN PEDRO"),
}


def rdp(points, epsilon):
    """Ramer-Douglas-Peucker iterativo (sin recursion, para no reventar el
    limite de recursion de Python con anillos de miles de puntos casi
    colineales). Simplifica cada poligono municipal ~80% sin alterar su
    forma de forma perceptible a la escala de este mapa (departamento/
    municipio, no catastral)."""
    if len(points) < 3:
        return points
    keep = bytearray(len(points))
    keep[0] = keep[-1] = 1
    pila = [(0, len(points) - 1)]
    while pila:
        i0, i1 = pila.pop()
        if i1 <= i0 + 1:
            continue
        x1, y1 = points[i0]; x2, y2 = points[i1]
        dx, dy = x2 - x1, y2 - y1
        norma2 = dx * dx + dy * dy
        dmax = -1; idx = -1
        for i in range(i0 + 1, i1):
            x, y = points[i]
            if norma2 == 0:
                d = ((x - x1) ** 2 + (y - y1) ** 2) ** 0.5
            else:
                t = ((x - x1) * dx + (y - y1) * dy) / norma2
                px, py = x1 + t * dx, y1 + t * dy
                d = ((x - px) ** 2 + (y - py) ** 2) ** 0.5
            if d > dmax:
                dmax = d; idx = i
        if dmax > epsilon:
            keep[idx] = 1
            pila.append((i0, idx))
            pila.append((idx, i1))
    return [p for p, k in zip(points, keep) if k]

DMS_RE = re.compile(r"^(\d+)\D+(\d+)\D+([\d.,]+)")


def parse_dms(texto):
    if not isinstance(texto, str):
        return None
    m = DMS_RE.match(texto.strip())
    if not m:
        return None
    deg = float(m.group(1))
    mnt = float(m.group(2))
    sec = float(m.group(3).replace(",", "."))
    return deg + mnt / 60 + sec / 3600


wb = openpyxl.load_workbook(FUENTE, read_only=True, data_only=True)
ws = wb["CUENTAME"]
rows = ws.iter_rows(min_row=1, values_only=True)
header = next(rows)
idx = {h: i for i, h in enumerate(header)}


def val(r, campo, default=None):
    v = r[idx[campo]]
    return v if v is not None else default


def txt(r, campo):
    v = r[idx[campo]]
    return str(v).strip() if v is not None else None


puntos = []
sin_coord = 0
for r in rows:
    if r[idx["Estado de la UDS"]] != "ACTIVA":
        continue
    lat = parse_dms(r[idx["Latitud UDS"]])
    lon = parse_dms(r[idx["Longitud UDS"]])
    if lat is None or lon is None:
        sin_coord += 1
        continue
    lon = -abs(lon)  # Antioquia siempre es Oeste; algunas filas no traen la letra O/W

    cupos = val(r, "Cupos UDS", 0) or 0
    atendidos = val(r, "Número de beneficiarios activos UDS", 0) or 0
    municipio = txt(r, "Municipio UDS") or ""

    puntos.append({
        "id": txt(r, "Código UDS"),
        "n": txt(r, "Unidad De Servicio (UDS)") or "",
        "en": txt(r, "Nombre Entidad Contratista") or "",
        "cz": txt(r, "Centro Zonal UDS") or "",
        "mun": municipio,
        "co": txt(r, "Comuna UDS"),
        "b": txt(r, "Barrio UDS"),
        "dir": txt(r, "Dirección UDS"),
        "tel": txt(r, "Teléfono UDS"),
        "cu": cupos,
        "at": atendidos,
        "clas": txt(r, "CLASIFICACIÓN COBERTURA"),
        "y": lat,
        "x": lon,
    })

# ---------- poligonos reales de los municipios ----------
# "antioquia_con_comunas v5.geojson": 146 features = 124 municipios (uno por
# poligono, propiedad MPIO_NOMBR) + las 21/22 comunas de Medellin (sin
# MPIO_NOMBR, ahi Medellin queda subdividido en vez de como un solo
# poligono) -- por eso Medellin no tiene "poligono" abajo y sigue usando
# solo el bbox derivado de sus propios puntos, igual que antes.
# Se cruza por la propiedad "name" (no "MPIO_NOMBR"): "MPIO_NOMBR" trae el
# nombre oficial completo (ej. "SAN PEDRO DE LOS MILAGROS") y en un caso
# viene directamente mal cargada en el geojson de origen (el feature de
# Zaragoza, COD_MPIO 05895, tiene MPIO_NOMBR="rionegro"); "name" siempre
# trae el nombre corto/correcto que coincide con "Municipio UDS" del Excel.
EPS_SIMPLIFICACION = 0.001  # grados (~110 m) -- imperceptible a esta escala, corta ~85% de los vertices
poligonos_por_muni = {}
medellin_lons, medellin_lats = [], []  # extent real de Medellin, desde sus 22 comunas (ver mas abajo)
with open(FUENTE_GEO, encoding="utf-8") as f:
    geo = json.load(f)
for ft in geo["features"]:
    props = ft["properties"]
    geom = ft["geometry"]
    anillos_src = geom["coordinates"] if geom["type"] == "Polygon" else [r for poly in geom["coordinates"] for r in poly]
    if not props.get("MPIO_NOMBR"):
        # es una comuna de Medellin (sin poligono municipal propio en esta
        # fuente): se usa solo para el extent real de la ciudad (ver abajo),
        # no se guarda como poligono navegable.
        for anillo in anillos_src:
            for lon, lat in anillo:
                medellin_lons.append(lon); medellin_lats.append(lat)
        continue
    nombre_geo = props.get("name") or props["MPIO_NOMBR"]
    key = norm(nombre_geo)
    key = ALIAS_MUNICIPIO.get(key, key)
    anillos = [[[lat, lon] for lon, lat in rdp([tuple(pt) for pt in anillo], EPS_SIMPLIFICACION)] for anillo in anillos_src]
    poligonos_por_muni[key] = {
        "anillos": anillos,
        "subregion": (props.get("SUBREGION") or "").strip(),
        "cod_mpio": props.get("COD_MPIO"),
    }

# bbox por municipio: del poligono real cuando hay (mas preciso que la nube
# de puntos, sobre todo en municipios grandes con UDS agrupadas en una sola
# esquina); para Medellin (unico sin poligono municipal en esta fuente) se
# usa el extent de sus 22 comunas en vez de sus propias UDS -- una UDS
# ("CI AGUAS FRIAS 2", codigo 050011159945) trae una coordenada mal
# digitada a 9.23N/-75.63 (deberia estar ~6.2N/-75.57, cerca de las demas),
# a mas de 300 km de Medellin real; usar el min/max de las UDS para el pin
# por defecto ponia el centro de la ciudad en un punto sin ninguna UDS
# cerca. El extent de las comunas (fuente geografica, no la captura de
# campo de esa UDS puntual) no tiene ese problema. La UDS en si se deja
# igual en D.puntos -- sigue siendo un resultado valido de busqueda, solo
# que nunca aparecera "cerca" de nada real dado donde quedo su coordenada.
# Margen ~5-15% para acotar el "viewbox" de Nominatim y el fitBounds inicial.

# ---------- casco urbano (cabecera municipal) de cada municipio ----------
# "COORDENADAS MUNICIPIOS.xlsx", hoja "bd def cntme": coordenada puntual de
# la cabecera de los 125 municipios (fuente propia del usuario, no derivada
# de las UDS) -- se usa para el pin por defecto en vez del centro geometrico
# del bbox/poligono, que puede caer en zona rural/deshabitada si el
# municipio tiene una forma alargada o irregular.
wb_c = openpyxl.load_workbook(FUENTE_CASCOS, read_only=True, data_only=True)
ws_c = wb_c["bd def cntme"]
rows_c = list(ws_c.iter_rows(min_row=1, values_only=True))
h_c = rows_c[0]; idx_c = {x: i for i, x in enumerate(h_c)}
casco_por_muni = {}
for r in rows_c[1:]:
    nom = r[idx_c["Nombre Municipio"]]
    lat_c, lon_c = r[idx_c["LATITUD"]], r[idx_c["LONGITUD"]]
    if nom and lat_c is not None and lon_c is not None:
        casco_por_muni[norm(nom)] = [lat_c, lon_c]

municipios = {}
for p in puntos:
    m = municipios.setdefault(p["mun"], {"nombre": p["mun"], "n": 0,
                                          "minx": p["x"], "maxx": p["x"],
                                          "miny": p["y"], "maxy": p["y"]})
    m["n"] += 1
    m["minx"] = min(m["minx"], p["x"]); m["maxx"] = max(m["maxx"], p["x"])
    m["miny"] = min(m["miny"], p["y"]); m["maxy"] = max(m["maxy"], p["y"])

n_con_poligono = 0
n_con_casco = 0
municipios_out = []
for m in municipios.values():
    poligono = poligonos_por_muni.get(norm(m["nombre"]))
    if poligono:
        n_con_poligono += 1
        lats = [pt[0] for anillo in poligono["anillos"] for pt in anillo]
        lons = [pt[1] for anillo in poligono["anillos"] for pt in anillo]
        minx, maxx, miny, maxy = min(lons), max(lons), min(lats), max(lats)
    elif norm(m["nombre"]) == norm("MEDELLIN") and medellin_lons:
        minx, maxx, miny, maxy = min(medellin_lons), max(medellin_lons), min(medellin_lats), max(medellin_lats)
    else:
        minx, maxx, miny, maxy = m["minx"], m["maxx"], m["miny"], m["maxy"]
    anchox = maxx - minx
    anchoy = maxy - miny
    padx = max(anchox * 0.05, 0.01)
    pady = max(anchoy * 0.05, 0.01)
    casco = casco_por_muni.get(norm(m["nombre"]))
    if casco:
        n_con_casco += 1
    municipios_out.append({
        "nombre": m["nombre"],
        "n": m["n"],
        "subregion": poligono["subregion"] if poligono else None,
        # bbox = [oeste, sur, este, norte], con margen
        "bbox": [minx - padx, miny - pady, maxx + padx, maxy + pady],
        # pin por defecto: casco urbano (cabecera municipal, COORDENADAS
        # MUNICIPIOS.xlsx) cuando existe -- fallback al centro geometrico
        # del poligono/bbox solo si algun municipio llegara a faltar ahi.
        "centro": casco or [(miny + maxy) / 2, (minx + maxx) / 2],
        "poligono": poligono["anillos"] if poligono else None,
    })
municipios_out.sort(key=lambda m: m["nombre"])
print(f"Municipios con poligono real: {n_con_poligono} / {len(municipios_out)}")
print(f"Municipios con casco urbano: {n_con_casco} / {len(municipios_out)}")

# bbox de todo el departamento, para la vista inicial antes de elegir municipio
todo_x = [p["x"] for p in puntos]
todo_y = [p["y"] for p in puntos]
bbox_depto = [min(todo_x), min(todo_y), max(todo_x), max(todo_y)]

data = {
    "meta": {
        "fuente": "BD ANALISIS COBERTURA v21.xlsm (Cuéntame ICBF), hoja CUENTAME, Regional Antioquia",
        "n_uds": len(puntos),
        "n_sin_coordenada": sin_coord,
        "n_municipios": len(municipios_out),
        "n_con_poligono": n_con_poligono,
    },
    "build": datetime.datetime.now().strftime("%d-%m-%Y %H:%M"),
    "bbox_depto": bbox_depto,
    "municipios": municipios_out,
    "puntos": puntos,
}

os.makedirs(os.path.dirname(OUT), exist_ok=True)
datos_str = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
with open(OUT, "w", encoding="utf-8") as f:
    f.write(datos_str)

print(f"UDS activas con coordenada: {len(puntos)} (sin coordenada utilizable: {sin_coord})")
print(f"Municipios: {len(municipios_out)}")
print(f"JSON: {len(datos_str) // 1024} KB -> {OUT}")
